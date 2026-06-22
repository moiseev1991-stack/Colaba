"""Excel-экспорт «На обзвон» для bulk-партии КП (2026-06-22).

Когда у компании из партии нет email-а — отправка КП по почте невозможна.
Но если есть телефон, юзер может «дотянуть» эту компанию руками: позвонить
или написать в WhatsApp. Этот модуль готовит .xlsx-файл со списком таких
компаний + контекст для звонка (боль из reviews, тема КП, полное тело).

Использует ту же `list_job_items` что и страница партии, фильтрует
строки без email + с валидным телефоном, и отдаёт xlsx через openpyxl
(требование уже есть в requirements.txt, см. website_leads_export.py).
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.outreach.kp_bulk_service import JobItemRow, list_job_items
from app.modules.outreach.phone_utils import (
    format_phone_for_display,
    is_russian_mobile,
    normalize_phone,
)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="334155")  # slate-700
_MOBILE_FILL = PatternFill("solid", fgColor="D1FAE5")  # emerald-100 — мобильный (WA)
_LANDLINE_FILL = PatternFill("solid", fgColor="F1F5F9")  # slate-100 — городской


def _wa_link(digits: str | None) -> str:
    """wa.me/{digits} без pre-fill — текст пишем юзером руками на старте звонка."""
    if not digits:
        return ""
    return f"https://wa.me/{digits}"


def _extract_pain_label(draft) -> str:
    """draft.arguments_used JSONB → pain_label (если был). Может отсутствовать
    у компаний без проанализированных отзывов."""
    if draft is None:
        return ""
    args = getattr(draft, "arguments_used", None)
    if not isinstance(args, dict):
        return ""
    return str(args.get("pain_label") or "")


def _extract_quote(draft) -> str:
    if draft is None:
        return ""
    args = getattr(draft, "arguments_used", None)
    if not isinstance(args, dict):
        return ""
    quote = args.get("quote")
    if not quote:
        return ""
    text = str(quote).strip().replace("\n", " ")
    if len(text) > 200:
        text = text[:200].rstrip() + "…"
    return text


_COLUMNS: list[tuple[str, int, str | None]] = [
    ("#", 5, None),
    ("Компания", 36, None),
    ("Город", 18, None),
    ("Телефон", 22, None),
    ("Тип", 12, None),
    ("WhatsApp ссылка", 36, None),
    ("Боль", 28, "wrap"),
    ("Цитата (для зацепки в разговоре)", 50, "wrap"),
    ("Что предложить (тема КП)", 50, "wrap"),
    ("Полное тело КП", 70, "wrap"),
]


def _row_for_xlsx(idx: int, item: JobItemRow, digits: str) -> list:
    is_mobile = is_russian_mobile(digits)
    return [
        idx,
        item.company_name or f"Компания #{item.company_id or ''}",
        item.company_city or "",
        format_phone_for_display(digits),
        "Мобильный" if is_mobile else "Городской",
        _wa_link(digits) if is_mobile else "",
        _extract_pain_label(item.draft),
        _extract_quote(item.draft),
        (item.draft.subject if item.draft else "") or "",
        (item.draft.body if item.draft else "") or "",
    ]


def _filter_callable_rows(items: list[JobItemRow]) -> list[tuple[JobItemRow, str]]:
    """Только компании без email, но с нормализуемым телефоном — это те
    кого юзер не может достать письмом, но может звонком/WA."""
    out: list[tuple[JobItemRow, str]] = []
    for it in items:
        if it.recipient_email:
            continue
        digits = normalize_phone(it.company_phone)
        if not digits:
            continue
        out.append((it, digits))
    return out


async def build_call_list_xlsx(
    db: AsyncSession, *, user_id: int, job_id: int
) -> tuple[bytes, int]:
    """Главная entry-point. Возвращает (bytes-файл, кол-во строк).

    Кол-во строк позволяет роутеру вернуть 204 No Content или 404, если
    звонить некому — иначе юзер качал бы пустой xlsx и недоумевал.
    """
    result = await list_job_items(db, user_id=user_id, job_id=job_id)
    if result is None:
        return b"", -1  # job not found / not owned

    _job, items = result
    callable_rows = _filter_callable_rows(items)

    wb = Workbook()
    ws = wb.active
    ws.title = "На обзвон"

    for col_idx, (header, width, _fmt) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_idx, (item, digits) in enumerate(callable_rows, start=2):
        is_mobile = is_russian_mobile(digits)
        values = _row_for_xlsx(row_idx - 1, item, digits)
        for col_idx, ((_header, _width, fmt), value) in enumerate(
            zip(_COLUMNS, values), start=1
        ):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fmt == "wrap":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Подсветка типа: мобильный — зелёным, городской — серым.
            # Помогает юзеру сразу видеть, кому можно WA, а кому только звонок.
            if col_idx == 5:  # колонка «Тип»
                cell.fill = _MOBILE_FILL if is_mobile else _LANDLINE_FILL
                cell.alignment = Alignment(vertical="center", horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(callable_rows)


def build_call_list_filename(job_id: int) -> str:
    """Имя файла для Content-Disposition. Включает id партии + дату скачивания,
    чтобы юзер не путал разные выгрузки в Downloads."""
    date = datetime.utcnow().strftime("%Y-%m-%d")
    return f"kp-call-list_job-{job_id}_{date}.xlsx"
