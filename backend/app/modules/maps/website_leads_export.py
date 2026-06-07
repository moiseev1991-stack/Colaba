"""Excel-экспорт website-лидов (блок 4 ТЗ 2026-06-02).

Генерирует .xlsx с двумя вкладками:
- «Лиды» — для отдела продаж: контакты + ИНН + website-score + драфт письма.
- «Производство сайта» — для верстальщика/генератора сайтов: AI-описание,
  адрес/координаты, часы работы, фото URLs, логотип URL, топ-цитаты, etc.

Использует openpyxl (см. requirements.txt). Если ассистент AI-описания
ещё не подключён — соответствующая колонка остаётся пустой (TODO в части C).

Возвращает байты .xlsx-файла; вызывающий endpoint оборачивает в
StreamingResponse с правильным Content-Disposition.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.lead_list import LeadList  # noqa: F401  (для будущего расширения)
from app.models.maps import Company, MapSearch, MapSearchResult, Review
from app.models.company_outreach_draft import CompanyOutreachDraft
from app.models.company_legal import CompanyLegal
from app.models.company_decision_maker import CompanyDecisionMaker
from app.models.pain_tag import CompanyPainScore, PainTag
from app.modules.maps import service as maps_service


logger = logging.getLogger(__name__)


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="334155")  # slate-700
_SCORE_HOT_FILL = PatternFill("solid", fgColor="DC2626")    # red-600
_SCORE_WARM_FILL = PatternFill("solid", fgColor="F59E0B")   # amber-500
_SCORE_COLD_FILL = PatternFill("solid", fgColor="94A3B8")   # slate-400


# Колонки вкладки «Лиды». (header, getter-callable, width, formatter)
def _build_leads_columns():
    return [
        ("Название", "name", 36, None),
        ("Город", "city", 18, None),
        ("Рубрика", "niche", 22, None),
        ("Телефон", "phone", 16, None),
        ("Email", "email_first", 28, None),
        ("Мессенджеры", "messengers", 22, None),
        ("Рейтинг", "rating", 8, None),
        ("Отзывов", "reviews_count", 9, None),
        ("Позитив", "reviews_positive_count", 9, None),
        ("Негатив", "reviews_negative_count", 9, None),
        ("Отвечает", "has_owner_replies", 10, None),
        ("Сайт", "website", 28, None),
        ("Источник", "source_label", 14, None),
        ("Pain-теги (топ-3)", "pain_tags_summary", 60, "wrap"),
        ("Адрес", "address", 36, None),
        ("ИНН", "legal_inn", 14, None),
        ("Юр.название", "legal_name", 32, None),
        ("Оборот", "legal_revenue", 14, None),
        ("Возраст (лет)", "legal_age", 12, None),
        ("ЛПР (ФИО)", "lpr_name", 28, None),
        ("ЛПР: должность", "lpr_post", 22, None),
        ("ЛПР: источник", "lpr_source_label", 18, None),
        ("Website-score", "website_lead_score", 12, "score"),
        ("Температура", "lead_temperature", 12, "score"),
        ("Ссылка на 2GIS/Я.Карты", "source_url", 40, None),
        ("Тема письма", "draft_subject", 40, None),
        ("Тело письма", "draft_body", 60, "wrap"),
    ]


def _build_production_columns():
    return [
        ("Название", "name", 36, None),
        ("Рубрика", "niche", 22, None),
        ("AI-описание (для hero/SEO)", "ai_description", 60, "wrap"),
        ("Город", "city", 18, None),
        ("Адрес", "address", 40, None),
        ("Lat", "lat", 12, None),
        ("Lng", "lng", 12, None),
        ("Часы работы", "working_hours", 32, "wrap"),
        ("Телефон", "phone", 16, None),
        ("Email", "email_first", 28, None),
        ("Мессенджеры", "messengers", 22, None),
        ("Рейтинг", "rating", 8, None),
        ("Отзывов", "reviews_count", 9, None),
        # ЛПР — для генератора hero «Связаться с директором Иваном Ивановичем»
        # (см. план в memory project_session_pending_2026-06-05.md).
        ("ЛПР (ФИО)", "lpr_name", 28, None),
        ("ЛПР: должность", "lpr_post", 22, None),
        ("Цитата 1", "quote_1", 60, "wrap"),
        ("Цитата 2", "quote_2", 60, "wrap"),
        ("Цитата 3", "quote_3", 60, "wrap"),
        ("Фото URLs", "photos", 60, "wrap"),
        ("Логотип URL", "logo_url", 60, None),
        ("Ссылка на источник", "source_url", 40, None),
    ]


def _score_fill(value: Any):
    if not isinstance(value, (int, float)):
        return None
    if value >= 70:
        return _SCORE_HOT_FILL
    if value >= 40:
        return _SCORE_WARM_FILL
    return _SCORE_COLD_FILL


def _build_source_deeplink(company: Company) -> str:
    src = (company.source or "").lower()
    ext = company.external_id or ""
    if not ext:
        return ""
    if src == "2gis":
        return f"https://2gis.ru/firm/{ext}"
    if src == "yandex_maps":
        return f"https://yandex.ru/maps/org/{ext}"
    return ""


def _extract_working_hours(c: Company) -> str:
    """Часы работы из raw_data провайдера (2GIS отдаёт schedule)."""
    raw = c.raw_data if isinstance(c.raw_data, dict) else None
    if not raw:
        return ""
    # 2GIS: items[].schedule
    sched = raw.get("schedule")
    if isinstance(sched, dict):
        parts = []
        for day, val in sched.items():
            if isinstance(val, dict):
                hours = val.get("working_hours") or val.get("hours")
                if isinstance(hours, list) and hours:
                    parts.append(f"{day}: " + ", ".join(
                        f"{h.get('from','?')}–{h.get('to','?')}" for h in hours
                    ))
        if parts:
            return "\n".join(parts)
    return ""


def _extract_photos(c: Company) -> list[str]:
    """Список URLs фото из raw_data (2GIS отдаёт photos.items)."""
    raw = c.raw_data if isinstance(c.raw_data, dict) else None
    if not raw:
        return []
    urls: list[str] = []
    photos = raw.get("photos") or raw.get("external_content")
    # 2GIS: photos.items[].main_photo_url
    if isinstance(photos, dict):
        items = photos.get("items") or photos.get("data") or []
        if isinstance(items, list):
            for it in items[:20]:
                if isinstance(it, dict):
                    url = (
                        it.get("main_photo_url")
                        or it.get("photo_url")
                        or it.get("url")
                    )
                    if isinstance(url, str) and url.startswith("http"):
                        urls.append(url)
    return urls


def _extract_logo_url(c: Company) -> str:
    """Лого: 2GIS отдаёт logo в external_content. Если нет — пусто."""
    raw = c.raw_data if isinstance(c.raw_data, dict) else None
    if not isinstance(raw, dict):
        return ""
    for key in ("logo", "logo_url", "icon"):
        v = raw.get(key)
        if isinstance(v, str) and v.startswith("http"):
            return v
    # Иногда логотип хранится в external_content.brand
    ec = raw.get("external_content")
    if isinstance(ec, dict):
        brand = ec.get("brand")
        if isinstance(brand, dict):
            url = brand.get("logo_url") or brand.get("url")
            if isinstance(url, str) and url.startswith("http"):
                return url
    return ""


async def _load_companies_for_search(
    db: AsyncSession, search_id: int, only_website_leads: bool
) -> list[Company]:
    """Тянет все компании поиска. only_website_leads=True — только те, у
    кого website_lead_score IS NOT NULL (то есть нет собственного сайта)."""
    stmt = (
        select(Company)
        .join(MapSearchResult, MapSearchResult.company_id == Company.id)
        .where(MapSearchResult.map_search_id == search_id)
        .order_by(MapSearchResult.position.asc())
    )
    if only_website_leads:
        stmt = stmt.where(Company.website_lead_score.isnot(None))
    res = await db.execute(stmt)
    return list(res.scalars().all())


async def _load_drafts(
    db: AsyncSession, company_ids: list[int]
) -> dict[int, CompanyOutreachDraft]:
    """Кэшированные drafts писем по компаниям (любой angle, берём свежее)."""
    if not company_ids:
        return {}
    stmt = (
        select(CompanyOutreachDraft)
        .where(CompanyOutreachDraft.company_id.in_(company_ids))
        .order_by(
            CompanyOutreachDraft.company_id.asc(),
            CompanyOutreachDraft.updated_at.desc(),
        )
    )
    res = await db.execute(stmt)
    drafts: dict[int, CompanyOutreachDraft] = {}
    for d in res.scalars().all():
        if d.company_id not in drafts:
            drafts[d.company_id] = d
    return drafts


async def _load_legal(
    db: AsyncSession, company_ids: list[int]
) -> dict[int, CompanyLegal]:
    """Юр.данные по компаниям (блок 2 ТЗ)."""
    if not company_ids:
        return {}
    stmt = select(CompanyLegal).where(CompanyLegal.company_id.in_(company_ids))
    res = await db.execute(stmt)
    out: dict[int, CompanyLegal] = {}
    for row in res.scalars().all():
        out[int(row.company_id)] = row
    return out


async def _load_top_decision_makers(
    db: AsyncSession, company_ids: list[int]
) -> dict[int, CompanyDecisionMaker]:
    """Топ-1 ЛПР с сайта по каждой компании (PR #15, миграция 032).

    Берём is_decision_maker=True, сортируем по confidence DESC — если
    несколько лиц найдено на /team, выбираем самого уверенного.
    """
    if not company_ids:
        return {}
    stmt = (
        select(CompanyDecisionMaker)
        .where(CompanyDecisionMaker.company_id.in_(company_ids))
        .where(CompanyDecisionMaker.is_decision_maker.is_(True))
        .order_by(
            CompanyDecisionMaker.company_id.asc(),
            CompanyDecisionMaker.confidence.desc(),
        )
    )
    res = await db.execute(stmt)
    out: dict[int, CompanyDecisionMaker] = {}
    for row in res.scalars().all():
        cid = int(row.company_id)
        if cid not in out:
            out[cid] = row
    return out


_DM_SOURCE_LABELS = {
    "website_team": "Сайт (/team)",
    "website_about": "Сайт (/о-нас)",
    "website_contacts": "Сайт (/контакты)",
}


async def _load_top_pain_tags(
    db: AsyncSession, company_ids: list[int], limit_per_company: int = 3
) -> dict[int, list[tuple[str, int, str | None]]]:
    """Топ N pain-теги по каждой компании: (label, mention_count, top_quote).

    Сортировка по mention_count DESC внутри компании. Если у компании
    мало болей — возвращаем меньше. Если совсем нет — нет записи в dict.
    """
    if not company_ids:
        return {}
    stmt = (
        select(
            CompanyPainScore.company_id,
            CompanyPainScore.mention_count,
            CompanyPainScore.top_quote,
            PainTag.label,
        )
        .join(PainTag, PainTag.id == CompanyPainScore.pain_tag_id)
        .where(CompanyPainScore.company_id.in_(company_ids))
        .where(CompanyPainScore.mention_count > 0)
        .order_by(
            CompanyPainScore.company_id.asc(),
            CompanyPainScore.mention_count.desc(),
        )
    )
    rows = list((await db.execute(stmt)).all())
    out: dict[int, list[tuple[str, int, str | None]]] = {}
    for cid, count, quote, label in rows:
        cid_int = int(cid)
        bucket = out.setdefault(cid_int, [])
        if len(bucket) >= limit_per_company:
            continue
        bucket.append((label, int(count), quote))
    return out


async def _load_company_sources_map(
    db: AsyncSession, company_ids: list[int]
) -> dict[int, set[str]]:
    """Множество источников по компании (2gis / yandex_maps / both)."""
    if not company_ids:
        return {}
    from sqlalchemy import text as sa_text

    sql = sa_text(
        "SELECT company_id, source FROM company_sources "
        "WHERE company_id = ANY(:ids)"
    )
    rows = list((await db.execute(sql, {"ids": list(company_ids)})).mappings().all())
    out: dict[int, set[str]] = {}
    for r in rows:
        out.setdefault(int(r["company_id"]), set()).add(r["source"])
    return out


_SOURCE_LABELS = {
    "2gis": "2GIS",
    "yandex_maps": "Я.Карты",
}


def _format_source_label(sources: set[str] | None, fallback_single: str | None) -> str:
    """Человекочитаемый источник: «2GIS», «Я.Карты», «Оба»."""
    if sources and len(sources) > 1:
        return "Оба"
    if sources:
        s = next(iter(sources))
        return _SOURCE_LABELS.get(s, s)
    if fallback_single:
        return _SOURCE_LABELS.get(fallback_single, fallback_single)
    return ""


def _format_pain_tags_summary(
    pains: list[tuple[str, int, str | None]] | None,
) -> str:
    """Текст для колонки «Pain-теги (топ-3)»: каждый тег с count и цитатой."""
    if not pains:
        return ""
    parts = []
    for label, count, quote in pains:
        line = f"• {label} (×{count})"
        if quote:
            quote_trimmed = quote.strip().replace("\n", " ")
            if len(quote_trimmed) > 140:
                quote_trimmed = quote_trimmed[:140].rstrip() + "…"
            line += f"\n  «{quote_trimmed}»"
        parts.append(line)
    return "\n".join(parts)


async def _load_top_positive_quotes(
    db: AsyncSession, company_ids: list[int], limit_per_company: int = 3
) -> dict[int, list[str]]:
    """Топ позитивных цитат по компаниям — для блока «Отзывы» на сайте.

    Берём positive-отзывы с непустым raw_text, по убыванию длины (но не
    >280 chars), уникальные.
    """
    if not company_ids:
        return {}
    out: dict[int, list[str]] = {}
    for cid in company_ids:
        stmt = (
            select(Review.raw_text)
            .where(Review.company_id == cid)
            .where(Review.sentiment == "positive")
            .where(Review.raw_text.isnot(None))
            .order_by(Review.posted_at.desc().nullslast())
            .limit(20)
        )
        rows = (await db.execute(stmt)).scalars().all()
        clean: list[str] = []
        seen: set[str] = set()
        for r in rows:
            t = (r or "").strip().replace("\n", " ")
            if not t or len(t) < 20:
                continue
            t = t[:280]
            if t in seen:
                continue
            seen.add(t)
            clean.append(t)
            if len(clean) >= limit_per_company:
                break
        if clean:
            out[cid] = clean
    return out


def _row_value(field: str, c: Company, ctx: dict) -> Any:
    """Достаёт значение для колонки. ctx содержит дополнительные данные
    (draft / quotes / legal_*), которые нельзя прочесть из самой Company."""
    if field == "email_first":
        emails = c.emails if isinstance(c.emails, list) else []
        return emails[0] if emails else ""
    if field == "messengers":
        extra = c.contacts_extra if isinstance(c.contacts_extra, dict) else {}
        parts = []
        for key, prefix in (
            ("telegrams", "tg"),
            ("whatsapps", "wa"),
            ("vks", "vk"),
        ):
            v = extra.get(key) if isinstance(extra, dict) else None
            if isinstance(v, list) and v:
                parts.append(f"{prefix}:{v[0]}")
        return ", ".join(parts)
    if field == "has_owner_replies":
        return "да" if c.has_owner_replies else "нет"
    if field == "rating":
        return float(c.rating) if c.rating is not None else None
    if field == "lat":
        return float(c.lat) if c.lat is not None else None
    if field == "lng":
        return float(c.lng) if c.lng is not None else None
    if field == "source_url":
        return _build_source_deeplink(c)
    if field == "working_hours":
        return _extract_working_hours(c)
    if field == "photos":
        urls = _extract_photos(c)
        return "\n".join(urls)
    if field == "logo_url":
        return _extract_logo_url(c)
    if field == "draft_subject":
        d = ctx.get("draft")
        return d.subject if d else ""
    if field == "draft_body":
        d = ctx.get("draft")
        return d.body if d else ""
    if field == "ai_description":
        # Часть C ТЗ блока 4: AI-описание компании. Пока заглушка —
        # будет реализована в отдельном LLM-ассистенте.
        return ctx.get("ai_description") or ""
    if field in ("quote_1", "quote_2", "quote_3"):
        idx = int(field.split("_")[-1]) - 1
        quotes = ctx.get("quotes") or []
        return quotes[idx] if idx < len(quotes) else ""
    if field in ("legal_inn", "legal_name", "legal_revenue", "legal_age"):
        legal = ctx.get("legal")
        if legal is None or legal.status != "ok":
            return ""
        if field == "legal_inn":
            return legal.inn or ""
        if field == "legal_name":
            return legal.legal_short_name or legal.legal_name or ""
        if field == "legal_revenue":
            return float(legal.revenue) if legal.revenue is not None else ""
        if field == "legal_age":
            return legal.age_years if legal.age_years is not None else ""
    if field == "source_label":
        return _format_source_label(
            ctx.get("company_sources") if isinstance(ctx.get("company_sources"), set) else None,
            getattr(c, "source", None),
        )
    if field == "pain_tags_summary":
        return _format_pain_tags_summary(ctx.get("pain_tags"))
    if field == "website":
        return getattr(c, "website", None) or ""
    if field in ("lpr_name", "lpr_post", "lpr_source_label"):
        legal = ctx.get("legal")
        dm = ctx.get("decision_maker")
        if legal is not None and legal.status == "ok" and legal.director_name:
            if field == "lpr_name":
                return legal.director_name
            if field == "lpr_post":
                return legal.director_post or ""
            return "DaData"
        if dm is not None:
            if field == "lpr_name":
                return dm.name
            if field == "lpr_post":
                return dm.post or ""
            return _DM_SOURCE_LABELS.get(dm.source, dm.source)
        return ""
    return getattr(c, field, "") or ""


def _write_sheet(ws, columns: list[tuple], rows: list[tuple[Company, dict]]):
    """Пишет заголовки + строки в один лист с типичным форматированием."""
    for col_idx, (header, _field, width, _fmt) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_idx, (company, ctx) in enumerate(rows, start=2):
        for col_idx, (_header, field, _width, fmt) in enumerate(columns, start=1):
            value = _row_value(field, company, ctx)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fmt == "wrap":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fmt == "score":
                fill = _score_fill(value)
                if fill is not None:
                    cell.fill = fill
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(
                        vertical="center", horizontal="center"
                    )


async def build_website_leads_xlsx(
    db: AsyncSession,
    search_id: int,
    *,
    only_website_leads: bool = True,
) -> bytes:
    """Главная entry-point: тянет компании, drafts, цитаты и собирает xlsx.

    only_website_leads=True (дефолт) — только те, у кого score IS NOT NULL.
    False — все компании поиска (для общего экспорта без фокуса на website).
    """
    companies = await _load_companies_for_search(db, search_id, only_website_leads)
    if not companies:
        # Возвращаем пустую книгу с заголовками — фронт получит файл, но
        # юзер увидит пустоту и поймёт что фильтр пустой.
        wb = Workbook()
        wb.active.title = "Лиды"
        ws_l = wb.active
        ws_p = wb.create_sheet("Производство сайта")
        _write_sheet(ws_l, _build_leads_columns(), [])
        _write_sheet(ws_p, _build_production_columns(), [])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    ids = [c.id for c in companies]
    drafts = await _load_drafts(db, ids)
    quotes_map = await _load_top_positive_quotes(db, ids, limit_per_company=3)
    legal_map = await _load_legal(db, ids)
    dm_map = await _load_top_decision_makers(db, ids)
    pain_tags_map = await _load_top_pain_tags(db, ids, limit_per_company=3)
    sources_map = await _load_company_sources_map(db, ids)

    # Блок 4C: автотриггер — для компаний без ai_description ставим
    # Celery-таски в фоне. Excel юзеру отдаём сразу с тем что есть;
    # повторный экспорт через 2-3 минуты получит заполненные описания.
    try:
        from app.modules.maps.company_description import (
            find_company_ids_without_description,
        )
        missing = await find_company_ids_without_description(db, ids)
        if missing:
            from app.modules.maps.tasks import generate_company_description
            for cid in missing[:200]:  # ограничиваем чтобы не закидать очередь
                try:
                    generate_company_description.delay(cid)
                except Exception as e:
                    logger.warning(
                        "build_website_leads_xlsx: cannot enqueue desc for #%d: %s",
                        cid, e,
                    )
            logger.info(
                "build_website_leads_xlsx: enqueued %d description tasks",
                min(len(missing), 200),
            )
    except Exception:
        logger.exception("description auto-enqueue failed (non-fatal)")

    rows: list[tuple[Company, dict]] = []
    for c in companies:
        legal = legal_map.get(c.id)
        ctx = {
            "draft": drafts.get(c.id),
            "quotes": quotes_map.get(c.id, []),
            "ai_description": c.ai_description or "",
            "legal": legal,
            "decision_maker": dm_map.get(c.id),
            "pain_tags": pain_tags_map.get(c.id, []),
            "company_sources": sources_map.get(c.id),
        }
        rows.append((c, ctx))

    wb = Workbook()
    ws_l = wb.active
    ws_l.title = "Лиды"
    ws_p = wb.create_sheet("Производство сайта")
    _write_sheet(ws_l, _build_leads_columns(), rows)
    _write_sheet(ws_p, _build_production_columns(), rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_filename(search: MapSearch) -> str:
    niche = (search.niche or "leads").replace(" ", "-")
    city = (search.city or "").replace(" ", "-")
    date = datetime.utcnow().strftime("%Y-%m-%d")
    if city:
        return f"website-leads_{niche}_{city}_{date}.xlsx"
    return f"website-leads_{niche}_{date}.xlsx"
