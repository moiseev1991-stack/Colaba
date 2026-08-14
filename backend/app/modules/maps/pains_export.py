"""Excel-экспорт выдачи «Боли» (GET /maps/pains/companies/export).

Один лист «Боли»: компании, отобранные по конкретной боли (pain_key или
pain_tag_ids), с контактами, метриками негатива и «пруф»-цитатой. Строит
.xlsx через openpyxl; вызывающий endpoint оборачивает в StreamingResponse.

На вход — те же строки, что и у /maps/pains/companies:
    (Company, mention_count, top_quote).
"""

from __future__ import annotations

import io
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.maps import Company


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="334155")  # slate-700


# (заголовок, ширина, getter(company, mentions, quote) -> значение)
def _columns() -> list[tuple[str, int, Any]]:
    return [
        ("Название", 36, lambda c, m, q: c.name),
        ("Ниша", 22, lambda c, m, q: c.niche),
        ("Город", 18, lambda c, m, q: c.city),
        ("Адрес", 34, lambda c, m, q: c.address),
        ("Телефон", 16, lambda c, m, q: c.phone),
        ("Сайт", 30, lambda c, m, q: c.website),
        ("Рейтинг", 9, lambda c, m, q: c.rating),
        ("Отзывов", 9, lambda c, m, q: c.reviews_count),
        ("Негатив", 9, lambda c, m, q: c.reviews_negative_count),
        ("Упоминаний боли", 16, lambda c, m, q: int(m or 0)),
        ("Температура", 12, lambda c, m, q: c.lead_temperature),
        ("Цитата-пруф", 60, lambda c, m, q: q),
        ("Источник", 14, lambda c, m, q: c.source),
    ]


def build_pains_xlsx(
    rows: Iterable[tuple[Company, Any, Any]],
    pain_labels: list[str],
    niche: str | None,
    city: str | None,
) -> bytes:
    """Возвращает байты .xlsx с одним листом «Боли»."""
    columns = _columns()
    wb = Workbook()
    ws = wb.active
    ws.title = "Боли"

    # Строка-контекст: какая боль/ниша/город выгружены.
    ctx_bits = []
    if pain_labels:
        ctx_bits.append("Боли: " + ", ".join(pain_labels))
    if niche:
        ctx_bits.append(f"Ниша: {niche}")
    if city:
        ctx_bits.append(f"Город: {city}")
    if ctx_bits:
        ws.append([" · ".join(ctx_bits)])
        ws.append([])

    header_row_idx = ws.max_row + 1
    ws.append([c[0] for c in columns])
    for col_idx, (title, width, _getter) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for company, mentions, top_quote in rows:
        ws.append([getter(company, mentions, top_quote) for _t, _w, getter in columns])

    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
